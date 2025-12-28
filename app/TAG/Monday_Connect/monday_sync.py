"""
Monday.com to Excel Sync Module
Fetches data from Monday.com and updates Excel file with change tracking
"""

import requests
import pandas as pd
import os
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
from change_logger import ChangeLogger


class MondaySync:
    """Syncs data from Monday.com to Excel with change tracking"""

    API_URL = "https://api.monday.com/v2"

    # Mapping of Monday.com groups to Excel sheets
    GROUP_TO_SHEET_MAPPING = {
        "Sales_Horizon": "SAL D'OURO HORIZON (9)",
        "Sal D'Ouro_Coast": "SAL D'OURO COAST (10)"
    }

    # Mapping of Monday.com column IDs to Excel column names
    COLUMN_MAPPING = {
        "text_mkr3f280": "Unit",
        "text_mkqy21fb": "Fraction",
        "text_mkqye2hw": "Layout",
        "numeric_mkqygjxw": "Floor",
        "color_mkqy1ck5": "Status",
        "color_mkr3pcv7": "Brokers company",
        "date_mkqymjsv": "Date of CPCV",
        "color_mkqyknad": "Client Nationality",
        "email_mkqw40b4": "Email",
        "phone_mkqzfwmj": "Phone"
    }

    # Board IDs
    DATA_BASE_CLIENTS_BOARD_ID = "1964802890"
    COMMISSIONS_MAP_BOARD_ID = "2021001619"

    def __init__(self, api_token: str, excel_path: str):
        """
        Initialize the sync module

        Args:
            api_token: Monday.com API token
            excel_path: Path to the Sales.xlsx file
        """
        self.api_token = api_token
        self.excel_path = excel_path
        self.headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
            "API-Version": "2024-01"
        }
        self.logger = ChangeLogger(os.path.dirname(excel_path))

    def _execute_query(self, query: str, variables: Optional[Dict] = None) -> Dict:
        """Execute a GraphQL query against Monday.com API"""
        payload = {"query": query}
        if variables:
            payload["variables"] = variables

        response = requests.post(
            self.API_URL,
            json=payload,
            headers=self.headers
        )
        response.raise_for_status()
        return response.json()

    def fetch_data_base_clients(self) -> List[Dict[str, Any]]:
        """Fetch all items from Data Base_Clients board"""
        query = """
        query {
            boards(ids: ["%s"]) {
                items_page(limit: 500) {
                    items {
                        id
                        name
                        group {
                            id
                            title
                        }
                        column_values {
                            id
                            text
                        }
                    }
                }
            }
        }
        """ % self.DATA_BASE_CLIENTS_BOARD_ID

        result = self._execute_query(query)

        if "data" not in result:
            raise Exception(f"API Error: {result.get('errors', 'Unknown error')}")

        items = result["data"]["boards"][0]["items_page"]["items"]

        # Parse items into structured format
        parsed_items = []
        for item in items:
            parsed = {
                "monday_id": item["id"],
                "client_name": item["name"],
                "group": item["group"]["title"]
            }

            for col in item["column_values"]:
                col_id = col["id"]
                if col_id in self.COLUMN_MAPPING:
                    parsed[self.COLUMN_MAPPING[col_id]] = col["text"] or ""

            parsed_items.append(parsed)

        return parsed_items

    def fetch_commissions_map(self) -> List[Dict[str, Any]]:
        """Fetch all items from Commissions Map board"""
        query = """
        query {
            boards(ids: ["%s"]) {
                items_page(limit: 500) {
                    items {
                        id
                        name
                        group {
                            id
                            title
                        }
                        column_values {
                            id
                            text
                        }
                    }
                }
            }
        }
        """ % self.COMMISSIONS_MAP_BOARD_ID

        result = self._execute_query(query)

        if "data" not in result:
            raise Exception(f"API Error: {result.get('errors', 'Unknown error')}")

        items = result["data"]["boards"][0]["items_page"]["items"]

        # Parse items
        parsed_items = []
        for item in items:
            cols = {c["id"]: c["text"] for c in item["column_values"] if c["text"]}
            parsed = {
                "monday_id": item["id"],
                "broker_name": item["name"],
                "group": item["group"]["title"],
                "project": cols.get("text_mks8b2xn", ""),
                "unit": cols.get("text_mks8krjd", ""),
                "client": cols.get("text_mks8pe8d", ""),
                "cpcv_date": cols.get("date_mks8j9vk", ""),
                "price": cols.get("numeric_mks8s09n", ""),
                "commission": cols.get("numeric_mks8pzrt", "")
            }
            parsed_items.append(parsed)

        return parsed_items

    def load_excel_sheet(self, sheet_name: str) -> pd.DataFrame:
        """Load a specific sheet from Excel"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
            df.columns = [str(c).replace('\n', ' ').strip() for c in df.columns]
            return df
        except Exception as e:
            raise Exception(f"Error loading sheet {sheet_name}: {e}")

    def find_matching_row(self, df: pd.DataFrame, monday_item: Dict[str, Any]) -> Optional[int]:
        """
        Find the matching row in Excel for a Monday.com item

        Matching strategy:
        1. Match by Fraction (unique identifier within a project)
        2. Or match by Unit number if Fraction not available
        """
        fraction = monday_item.get("Fraction", "")
        unit = monday_item.get("Unit", "")

        if fraction and "Fraction" in df.columns:
            matches = df[df["Fraction"].astype(str).str.strip() == fraction.strip()]
            if len(matches) == 1:
                return matches.index[0]

        if unit and "Unit" in df.columns:
            matches = df[df["Unit"].astype(str).str.strip() == unit.strip()]
            if len(matches) == 1:
                return matches.index[0]

        return None

    def compare_and_detect_changes(
        self,
        excel_row: pd.Series,
        monday_item: Dict[str, Any],
        fields_to_compare: List[str]
    ) -> List[Dict[str, Any]]:
        """
        Compare Excel row with Monday.com item and detect changes

        Returns list of changes with field name, old value, new value

        Note: Only updates Excel if Monday.com has a value. Does NOT
        overwrite Excel data with empty Monday.com values.
        """
        changes = []

        for field in fields_to_compare:
            if field not in monday_item:
                continue

            monday_value = str(monday_item.get(field, "")).strip()
            excel_value = str(excel_row.get(field, "")).strip() if field in excel_row.index else ""

            # Normalize values for comparison
            if pd.isna(excel_row.get(field)) or excel_row.get(field) == "":
                excel_value = ""

            # Skip if both are empty
            if not monday_value and not excel_value:
                continue

            # IMPORTANT: Only update if Monday.com has data
            # Don't overwrite Excel values with empty Monday.com values
            if not monday_value and excel_value:
                continue

            # Compare values
            if monday_value != excel_value:
                changes.append({
                    "field": field,
                    "old_value": excel_value,
                    "new_value": monday_value
                })

        return changes

    def create_backup(self) -> str:
        """Create a backup of the Excel file before modifications"""
        import shutil
        from datetime import datetime

        backup_dir = os.path.join(os.path.dirname(self.excel_path), "backups")
        os.makedirs(backup_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"Sales_backup_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_filename)

        shutil.copy2(self.excel_path, backup_path)
        return backup_path

    def update_excel_cell(self, sheet_name: str, row: int, col_name: str, value: Any):
        """
        Update a single cell in Excel using openpyxl

        WARNING: openpyxl may not preserve all Excel features (Data Validation, etc.)
        A backup is created before any modifications.
        """
        # ALWAYS create backup before first modification
        if not hasattr(self, '_backup_created') or not self._backup_created:
            backup_path = self.create_backup()
            self._backup_created = True
            print(f"Backup created: {backup_path}")

        wb = load_workbook(self.excel_path)
        ws = wb[sheet_name]

        # Find column index (1-based for openpyxl)
        header_row = [cell.value for cell in ws[1]]
        # Clean header names
        header_row = [str(h).replace('\n', ' ').strip() if h else "" for h in header_row]

        try:
            col_idx = header_row.index(col_name) + 1
        except ValueError:
            wb.close()
            raise Exception(f"Column '{col_name}' not found in sheet '{sheet_name}'")

        # Update cell (row is 0-indexed from pandas, +2 for 1-indexed + header)
        ws.cell(row=row + 2, column=col_idx, value=value)

        wb.save(self.excel_path)
        wb.close()

    def sync_data_base_clients(self, dry_run: bool = False) -> Dict[str, Any]:
        """
        Sync Data Base_Clients board to Excel sheets

        Args:
            dry_run: If True, only detect changes without updating Excel

        Returns:
            Sync summary with changes detected/applied
        """
        self.logger.start_sync()
        changes_summary = {
            "sheets_processed": [],
            "total_items_checked": 0,
            "total_changes": 0,
            "errors": []
        }

        try:
            # Fetch data from Monday.com
            monday_items = self.fetch_data_base_clients()
            changes_summary["total_items_checked"] = len(monday_items)

            # Group items by their target sheet
            items_by_sheet = {}
            for item in monday_items:
                group = item.get("group", "")
                sheet_name = self.GROUP_TO_SHEET_MAPPING.get(group)
                if sheet_name:
                    if sheet_name not in items_by_sheet:
                        items_by_sheet[sheet_name] = []
                    items_by_sheet[sheet_name].append(item)

            # Process each sheet
            fields_to_sync = ["Status", "Brokers company", "Date of CPCV", "Client Nationality"]

            for sheet_name, items in items_by_sheet.items():
                try:
                    # Load Excel sheet
                    df = self.load_excel_sheet(sheet_name)
                    changes_summary["sheets_processed"].append(sheet_name)

                    for item in items:
                        # Find matching row
                        row_idx = self.find_matching_row(df, item)

                        if row_idx is None:
                            # Item exists in Monday but not in Excel - could be new
                            continue

                        # Detect changes
                        excel_row = df.iloc[row_idx]
                        changes = self.compare_and_detect_changes(
                            excel_row, item, fields_to_sync
                        )

                        # Also check Client name
                        client_name = item.get("client_name", "")
                        excel_client = str(excel_row.get("Client", "")).strip() if "Client" in excel_row.index else ""
                        if client_name and client_name != excel_client:
                            changes.append({
                                "field": "Client",
                                "old_value": excel_client,
                                "new_value": client_name
                            })

                        # Log and apply changes
                        for change in changes:
                            self.logger.log_change(
                                source_board="Data Base_Clients",
                                target_sheet=sheet_name,
                                unit=item.get("Fraction", item.get("Unit", "")),
                                field=change["field"],
                                old_value=change["old_value"],
                                new_value=change["new_value"],
                                monday_item_id=item.get("monday_id", ""),
                                row_index=row_idx
                            )

                            if not dry_run:
                                try:
                                    self.update_excel_cell(
                                        sheet_name,
                                        row_idx,
                                        change["field"],
                                        change["new_value"]
                                    )
                                except Exception as e:
                                    changes_summary["errors"].append(
                                        f"Failed to update {sheet_name}[{row_idx}].{change['field']}: {e}"
                                    )

                            changes_summary["total_changes"] += 1

                except Exception as e:
                    changes_summary["errors"].append(f"Error processing sheet {sheet_name}: {e}")

            # End sync and save log
            summary = self.logger.end_sync(
                success=len(changes_summary["errors"]) == 0,
                error_message="; ".join(changes_summary["errors"]) if changes_summary["errors"] else ""
            )

            changes_summary["sync_id"] = summary["sync_id"]
            changes_summary["dry_run"] = dry_run

        except Exception as e:
            self.logger.end_sync(success=False, error_message=str(e))
            changes_summary["errors"].append(str(e))

        return changes_summary

    def preview_changes(self) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Preview changes without applying them

        Returns:
            Tuple of (changes_dataframe, summary_dict)
        """
        # Run sync in dry run mode
        summary = self.sync_data_base_clients(dry_run=True)

        # Get changes as dataframe
        changes_df = self.logger.get_changes_as_dataframe(limit=200)

        return changes_df, summary

    def get_sync_status(self) -> Dict[str, Any]:
        """Get current sync status and recent activity"""
        recent_syncs = self.logger.get_recent_syncs(limit=5)

        last_sync = recent_syncs[0] if recent_syncs else None

        return {
            "last_sync_time": last_sync["timestamp"] if last_sync else None,
            "last_sync_success": last_sync["success"] if last_sync else None,
            "last_sync_changes": last_sync["total_changes"] if last_sync else 0,
            "recent_syncs": len(recent_syncs),
            "total_logged_syncs": len(self.logger.load_all_logs())
        }


def create_sync_instance(api_token: str = None) -> MondaySync:
    """
    Factory function to create a MondaySync instance

    Args:
        api_token: Monday.com API token. If not provided, reads from environment.

    Returns:
        Configured MondaySync instance
    """
    if not api_token:
        api_token = os.getenv("MONDAY_API_TOKEN")

    if not api_token:
        raise ValueError("API token is required. Set MONDAY_API_TOKEN environment variable.")

    excel_path = os.path.join(os.path.dirname(__file__), "Sales.xlsx")

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    return MondaySync(api_token, excel_path)
